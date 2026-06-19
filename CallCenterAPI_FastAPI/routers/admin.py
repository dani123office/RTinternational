import bcrypt
import io
import traceback
from datetime import datetime, date
from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy import func
from ..database import get_db
from ..models import User, Customer, CallBack, Transfer, Sale, ActivityLog, Attendance, LeaveRequest, LoanRequest, Notification, EmailVerification, StaffOTP
from .auth import require_admin
from ..utils.logger import log_activity, get_client_ip
from ..utils.email import generate_otp, send_otp_email
from ..schemas import (
    CreateManagerRequest, CreateAgentRequest, AssignAgentRequest,
    UpdateUserRequest, UpdateAgentStaffRequest, ApproveUserRequest, ResetUserPasswordRequest,
    OverallStats, ManagerKpi, AgentKpi, AgentDetail, AgentStats, AgentOut,
    AdminPerformanceOverview, BusinessFeedItem, SendOTPRequest, VerifyOTPRequest,
)
from .callbacks import _build_callback_out
from .transfers import _transfer_out
from .sales import _sale_out
from .salary import _SalaryPDF, _employee_id, _month_name, _attendance_summary, _fmt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api/admin", tags=["admin"])


def _safe_div(a: int, b: int) -> float:
    if b == 0:
        return 0.0
    return round((a / b) * 100, 1)


# ─── User Management ─────────────────────────────────────


@router.post("/create-manager", status_code=201)
def create_manager(data: CreateManagerRequest, request: Request, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    try:
        email = data.email.strip().lower()
        existing = db.query(User).filter(func.lower(User.email) == email).first()
        if existing:
            raise HTTPException(status_code=400, detail="Email already exists")
        hashed = bcrypt.hashpw(data.password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        user = User(
            name=data.name,
            email=email,
            password_hash=hashed,
            role="manager",
            is_active=1,
        )
        db.add(user)
        db.flush()
        ev = db.query(EmailVerification).filter(EmailVerification.user_id == user.id).first()
        if not ev:
            db.add(EmailVerification(user_id=user.id, is_verified=1))
        db.commit()
        db.refresh(user)
        log_activity(db, admin.id, "created", "manager", user.id,
                     f"Created manager {user.email}",
                     get_client_ip(request))
        return {"id": user.id, "name": user.name, "email": user.email, "role": user.role}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Failed to create manager: {str(e)}")


@router.post("/create-agent", status_code=201)
def create_agent(data: CreateAgentRequest, request: Request, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    try:
        email = data.email.strip().lower()
        existing = db.query(User).filter(func.lower(User.email) == email).first()
        if existing:
            raise HTTPException(status_code=400, detail="Email already exists")
        manager = db.query(User).filter(
            User.id == data.managerId, User.role == "manager", User.is_active == True,
        ).first()
        if not manager:
            raise HTTPException(status_code=400, detail="Manager not found or inactive")
        hashed = bcrypt.hashpw(data.password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        user = User(
            name=data.name,
            email=email,
            password_hash=hashed,
            role="agent",
            manager_id=data.managerId,
            is_active=1,
            father_name=data.fatherName,
            monthly_salary=data.monthlySalary,
            cnic=data.cnic,
            phone=data.phone,
            department=data.department,
            designation=data.designation,
            date_of_birth=data.dateOfBirth,
            date_of_joining=data.dateOfJoining,
            emerg_contact_name=data.emergContactName,
            emerg_contact_number=data.emergContactNumber,
        )
        db.add(user)
        db.flush()
        # Ensure email verification is required — mark as unverified
        ev = db.query(EmailVerification).filter(EmailVerification.user_id == user.id).first()
        if not ev:
            ev = EmailVerification(user_id=user.id, is_verified=0)
            db.add(ev)
        db.commit()
        db.refresh(user)
        log_activity(db, admin.id, "created", "agent", user.id,
                     f"Created agent {user.email}",
                     get_client_ip(request))
        return {"id": user.id, "name": user.name, "email": user.email, "role": user.role, "managerId": user.manager_id}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Failed to create agent: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Failed to create agent: {str(e)}")


@router.get("/users")
def get_all_users(admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    users = db.query(User).order_by(User.role, User.name).all()
    return [
        {
            "id": u.id, "name": u.name, "email": u.email,
            "role": u.role, "isActive": 1 if u.is_active else 0,
            "managerId": u.manager_id,
            "phone": u.phone,
            "fatherName": u.father_name,
            "monthlySalary": u.monthly_salary,
            "cnic": u.cnic,
            "department": u.department,
            "designation": u.designation,
            "dateOfBirth": u.date_of_birth.isoformat() if u.date_of_birth else None,
            "dateOfJoining": u.date_of_joining.isoformat() if u.date_of_joining else None,
            "emergContactName": u.emerg_contact_name,
            "emergContactNumber": u.emerg_contact_number,
            "createdAt": u.created_at.isoformat() if u.created_at else None,
        }
        for u in users
    ]


@router.get("/managers")
def get_managers(
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    managers = db.query(User).filter(User.role == "manager", User.is_active == True).order_by(User.name).all()
    result = []
    for m in managers:
        agent_ids = [a.id for a in db.query(User).filter(User.manager_id == m.id).all()]
        if agent_ids:
            cb = db.query(func.count(CallBack.id)).filter(CallBack.employee_id.in_(agent_ids)).scalar() or 0
            tr = db.query(func.count(Transfer.id)).filter(Transfer.employee_id.in_(agent_ids)).scalar() or 0
            sa = db.query(func.count(Sale.id)).filter(Sale.employee_id.in_(agent_ids)).scalar() or 0
        else:
            cb = tr = sa = 0
        total_opps = tr
        result.append(ManagerKpi(
            id=m.id, name=m.name, teamSize=len(agent_ids),
            callbacks=cb, transfers=tr, sales=sa,
            conversionRate=_safe_div(sa, total_opps),
        ))
    return result


@router.get("/agents")
def get_agents(
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
    showAll: bool = False,
):
    query = db.query(User).filter(User.role == "agent")
    if not showAll:
        query = query.filter(User.is_active == True)
    agents = query.order_by(User.name).all()
    manager_map = {}
    for a in agents:
        if a.manager_id not in manager_map:
            mgr = db.query(User).filter(User.id == a.manager_id).first()
            manager_map[a.manager_id] = mgr.name if mgr else "Unassigned"
    result = []
    for a in agents:
        cb = db.query(func.count(CallBack.id)).filter(CallBack.employee_id == a.id).scalar() or 0
        tr = db.query(func.count(Transfer.id)).filter(Transfer.employee_id == a.id).scalar() or 0
        sa = db.query(func.count(Sale.id)).filter(Sale.employee_id == a.id).scalar() or 0
        total_opps = tr
        result.append(AgentKpi(
            id=a.id, name=a.name,
            managerName=manager_map.get(a.manager_id, "Unassigned"),
            callbacks=cb, transfers=tr, sales=sa,
            conversionRate=_safe_div(sa, total_opps),
            isActive=1 if a.is_active else 0,
            monthlySalary=a.monthly_salary if a.monthly_salary else 0,
        ))
    return result


@router.get("/managers/{manager_id}")
def get_manager_detail(
    manager_id: int,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    manager = db.query(User).filter(User.id == manager_id, User.role == "manager").first()
    if not manager:
        raise HTTPException(status_code=404, detail="Manager not found")

    agents = db.query(User).filter(User.manager_id == manager.id, User.role == "agent").all()

    agent_list = []
    for a in agents:
        cb = db.query(func.count(CallBack.id)).filter(CallBack.employee_id == a.id).scalar() or 0
        tr = db.query(func.count(Transfer.id)).filter(Transfer.employee_id == a.id).scalar() or 0
        sa = db.query(func.count(Sale.id)).filter(Sale.employee_id == a.id).scalar() or 0
        total_opps = tr
        agent_list.append(AgentKpi(
            id=a.id, name=a.name,
            managerName=manager.name,
            callbacks=cb, transfers=tr, sales=sa,
            conversionRate=_safe_div(sa, total_opps),
            isActive=1 if a.is_active else 0,
        ))

    return {
        "manager": {
            "id": manager.id, "name": manager.name, "email": manager.email,
            "role": manager.role, "isActive": 1 if manager.is_active else 0,
            "createdAt": manager.created_at.isoformat() if manager.created_at else None,
        },
        "agents": agent_list,
        "teamSize": len(agents),
    }


@router.get("/agents/{agent_id}")
def get_agent_detail(
    agent_id: int,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    agent = db.query(User).filter(User.id == agent_id, User.role == "agent").first()
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")

    callbacks = db.query(CallBack).filter(CallBack.employee_id == agent.id).all()
    transfers = db.query(Transfer).filter(Transfer.employee_id == agent.id).all()
    sales = db.query(Sale).filter(Sale.employee_id == agent.id).all()

    ac = len(callbacks)
    at = len(transfers)
    as_ = len(sales)
    a_total = ac + at
    a_cr = _safe_div(as_, a_total)

    return AgentDetail(
        agent=AgentOut(
            id=agent.id, name=agent.name, email=agent.email,
            role=agent.role, isActive=agent.is_active, managerId=agent.manager_id,
            phone=agent.phone,
            fatherName=agent.father_name,
            monthlySalary=agent.monthly_salary,
            cnic=agent.cnic,
            department=agent.department,
            designation=agent.designation,
            dateOfBirth=agent.date_of_birth,
            dateOfJoining=agent.date_of_joining,
            emergContactName=agent.emerg_contact_name,
            emergContactNumber=agent.emerg_contact_number,
        ),
        callbacks=[_build_callback_out(c, c.customer) for c in callbacks],
        transfers=[_transfer_out(t, t.customer) for t in transfers],
        sales=[_sale_out(s, s.customer) for s in sales],
        stats=AgentStats(
            agent=AgentOut(
                id=agent.id, name=agent.name, email=agent.email,
                role=agent.role, isActive=agent.is_active, managerId=agent.manager_id,
                phone=agent.phone,
                fatherName=agent.father_name,
                monthlySalary=agent.monthly_salary,
                cnic=agent.cnic,
                department=agent.department,
                designation=agent.designation,
                dateOfBirth=agent.date_of_birth,
                dateOfJoining=agent.date_of_joining,
                emergContactName=agent.emerg_contact_name,
                emergContactNumber=agent.emerg_contact_number,
            ),
            callbacks=ac, transfers=at, sales=as_, conversionRate=a_cr,
        ),
    )


@router.put("/agents/{agent_id}")
def update_agent_staff(
    agent_id: int,
    data: UpdateAgentStaffRequest,
    request: Request,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    try:
        agent = db.query(User).filter(User.id == agent_id, User.role == "agent").first()
        if not agent:
            raise HTTPException(status_code=404, detail="Agent not found")
        if data.name is not None:
            agent.name = data.name
        if data.email is not None:
            clean = data.email.strip().lower()
            existing = db.query(User).filter(func.lower(User.email) == clean, User.id != agent_id).first()
            if existing:
                raise HTTPException(status_code=400, detail="Email already in use")
            agent.email = clean
        if data.phone is not None:
            agent.phone = data.phone or None
        if data.fatherName is not None:
            agent.father_name = data.fatherName or None
        if data.monthlySalary is not None:
            agent.monthly_salary = data.monthlySalary
        if data.cnic is not None:
            agent.cnic = data.cnic or None
        if data.department is not None:
            agent.department = data.department or None
        if data.designation is not None:
            agent.designation = data.designation or None
        if data.dateOfBirth is not None:
            agent.date_of_birth = data.dateOfBirth
        if data.dateOfJoining is not None:
            agent.date_of_joining = data.dateOfJoining
        if data.emergContactName is not None:
            agent.emerg_contact_name = data.emergContactName or None
        if data.emergContactNumber is not None:
            agent.emerg_contact_number = data.emergContactNumber or None
        db.commit()
        db.refresh(agent)
        log_activity(db, admin.id, "updated", "agent", agent.id,
                     f"Updated agent #{agent.id} staff info",
                     get_client_ip(request))
        return AgentOut(
            id=agent.id, name=agent.name, email=agent.email,
            role=agent.role, isActive=agent.is_active, managerId=agent.manager_id,
            phone=agent.phone,
            fatherName=agent.father_name,
            monthlySalary=agent.monthly_salary,
            cnic=agent.cnic,
            department=agent.department,
            designation=agent.designation,
            dateOfBirth=agent.date_of_birth,
            dateOfJoining=agent.date_of_joining,
            emergContactName=agent.emerg_contact_name,
            emergContactNumber=agent.emerg_contact_number,
        )
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Failed to update agent: {str(e)}")


@router.put("/user/{user_id}")
def update_user(
    user_id: int, data: UpdateUserRequest,
    request: Request,
    admin: User = Depends(require_admin), db: Session = Depends(get_db),
):
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        if data.name is not None:
            user.name = data.name
        if data.email is not None:
            clean = data.email.strip().lower()
            existing = db.query(User).filter(func.lower(User.email) == clean, User.id != user_id).first()
            if existing:
                raise HTTPException(status_code=400, detail="Email already in use")
            user.email = clean
        if data.isActive is not None:
            user.is_active = data.isActive
        if data.managerId is not None:
            if user.role != "agent":
                raise HTTPException(status_code=400, detail="Only agents can be assigned a manager")
            mgr = db.query(User).filter(
                User.id == data.managerId, User.role == "manager", User.is_active == True,
            ).first()
            if not mgr:
                raise HTTPException(status_code=400, detail="Manager not found or inactive")
            user.manager_id = data.managerId
        db.commit()
        db.refresh(user)
        log_activity(db, admin.id, "updated", "user", user.id,
                     f"Updated user #{user.id}",
                     get_client_ip(request))
        return {"id": user.id, "name": user.name, "email": user.email, "role": user.role, "isActive": 1 if user.is_active else 0, "managerId": user.manager_id}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Failed to update user: {str(e)}")


@router.delete("/user/{user_id}")
def delete_user(user_id: int, request: Request, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        if user.role == "admin":
            raise HTTPException(status_code=400, detail="Cannot delete admin account")
        if user.role == "manager":
            agent_count = db.query(User).filter(User.manager_id == user.id).count()
            if agent_count > 0:
                raise HTTPException(status_code=400, detail="Unable to delete manager. First delete all agents associated with this manager")
        name = user.name
        # Delete activity logs
        db.query(ActivityLog).filter(ActivityLog.user_id == user.id).delete()
        # Delete callbacks where this user is the employee or creating manager
        db.query(CallBack).filter(CallBack.employee_id == user.id).delete()
        db.query(CallBack).filter(CallBack.created_by_manager_id == user.id).delete()
        # Delete transfers where this user is the employee
        db.query(Transfer).filter(Transfer.employee_id == user.id).delete()
        # Delete sales where this user is the employee
        db.query(Sale).filter(Sale.employee_id == user.id).delete()
        # Before deleting customers, remove all child records linked to them
        customer_ids = [c.id for c in db.query(Customer.id).filter(Customer.created_by == user.id).all()]
        if customer_ids:
            db.query(CallBack).filter(CallBack.customer_id.in_(customer_ids)).delete(synchronize_session=False)
            db.query(Transfer).filter(Transfer.customer_id.in_(customer_ids)).delete(synchronize_session=False)
            db.query(Sale).filter(Sale.customer_id.in_(customer_ids)).delete(synchronize_session=False)
        db.query(Customer).filter(Customer.created_by == user.id).delete()
        # Delete notification, attendance, leave, loan records for this user
        db.query(Notification).filter(Notification.user_id == user.id).delete()
        db.query(Attendance).filter(Attendance.user_id == user.id).delete()
        db.query(LeaveRequest).filter(LeaveRequest.user_id == user.id).delete()
        db.query(LeaveRequest).filter(LeaveRequest.admin_id == user.id).update({"admin_id": None})
        db.query(LoanRequest).filter(LoanRequest.user_id == user.id).delete()
        db.query(LoanRequest).filter(LoanRequest.admin_id == user.id).update({"admin_id": None})
        # Delete email verification records (must be before user deletion)
        db.query(EmailVerification).filter(EmailVerification.user_id == user.id).delete()
        u_id = user.id
        db.delete(user)
        db.commit()
        log_activity(db, admin.id, "deleted", "user", u_id,
                     f"Deleted user {name}",
                     get_client_ip(request))
        return {"ok": True, "message": f"User {name} and all associated records permanently deleted"}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        tb = traceback.format_exc()
        print(f"[DELETE_USER ERROR] user_id={user_id}: {e}\n{tb}")
        raise HTTPException(status_code=400, detail=f"Failed to delete user: {str(e)}")


@router.patch("/assign-agent")
def assign_agent(data: AssignAgentRequest, request: Request, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    try:
        agent = db.query(User).filter(User.id == data.agentId, User.role == "agent").first()
        if not agent:
            raise HTTPException(status_code=404, detail="Agent not found")
        manager = db.query(User).filter(
            User.id == data.managerId, User.role == "manager", User.is_active == True,
        ).first()
        if not manager:
            raise HTTPException(status_code=400, detail="Manager not found or inactive")
        agent.manager_id = data.managerId
        db.commit()
        log_activity(db, admin.id, "updated", "agent_assignment", agent.id,
                     f"Assigned agent #{agent.id} to manager #{data.managerId}",
                     get_client_ip(request))
        return {"ok": True, "agentId": agent.id, "managerId": manager.id}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Failed to assign agent: {str(e)}")


# ─── Pending Users (Admin Approval) ──────────────────────


@router.get("/pending-users")
def get_pending_users(admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    users = db.query(User).filter(
        User.is_active == False,
        User.role.in_(["agent", "manager"]),
    ).order_by(User.created_at.desc()).all()
    return [
        {
            "id": u.id, "name": u.name, "email": u.email,
            "role": u.role, "createdAt": u.created_at.isoformat() if u.created_at else None,
            "phone": u.phone,
            "fatherName": u.father_name,
            "monthlySalary": u.monthly_salary,
            "cnic": u.cnic,
            "department": u.department,
            "designation": u.designation,
            "dateOfBirth": u.date_of_birth.isoformat() if u.date_of_birth else None,
            "dateOfJoining": u.date_of_joining.isoformat() if u.date_of_joining else None,
            "emergContactName": u.emerg_contact_name,
            "emergContactNumber": u.emerg_contact_number,
        }
        for u in users
    ]


@router.post("/approve-user/{user_id}")
def approve_user(
    user_id: int, data: ApproveUserRequest,
    request: Request,
    admin: User = Depends(require_admin), db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.is_active:
        raise HTTPException(status_code=400, detail="User is already active")

    manager = db.query(User).filter(
        User.id == data.managerId, User.role == "manager", User.is_active == True,
    ).first()
    if not manager:
        raise HTTPException(status_code=400, detail="Manager not found or inactive")

    if data.name is not None:
        user.name = data.name
    if data.email is not None:
        user.email = data.email.strip().lower()
    if data.phone is not None:
        user.phone = data.phone
    if data.fatherName is not None:
        user.father_name = data.fatherName
    if data.monthlySalary is not None:
        user.monthly_salary = data.monthlySalary
    if data.cnic is not None:
        user.cnic = data.cnic
    if data.department is not None:
        user.department = data.department
    if data.designation is not None:
        user.designation = data.designation
    if data.dateOfBirth is not None:
        user.date_of_birth = data.dateOfBirth
    if data.dateOfJoining is not None:
        user.date_of_joining = data.dateOfJoining
    if data.emergContactName is not None:
        user.emerg_contact_name = data.emergContactName
    if data.emergContactNumber is not None:
        user.emerg_contact_number = data.emergContactNumber

    user.manager_id = data.managerId
    user.is_active = True
    db.commit()
    db.refresh(user)
    log_activity(db, admin.id, "approved", "user", user.id,
                 f"Approved user {user.email}",
                 get_client_ip(request))
    return {"ok": True, "message": f"User {user.name} approved and assigned to {manager.name}"}


@router.put("/users/{user_id}/reset-password")
def reset_user_password(
    user_id: int, data: ResetUserPasswordRequest,
    request: Request,
    admin: User = Depends(require_admin), db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if len(data.newPassword) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    if data.oldPassword:
        if not bcrypt.checkpw(data.oldPassword.encode("utf-8"), user.password_hash.encode("utf-8")):
            raise HTTPException(status_code=400, detail="Old password is incorrect")
    user.password_hash = bcrypt.hashpw(data.newPassword.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    db.commit()
    log_activity(db, admin.id, "password_reset", "user", user.id,
                 f"Password reset for user {user.email}",
                 get_client_ip(request))
    return {"ok": True, "message": f"Password for {user.name} has been reset"}


# ─── Analytics ────────────────────────────────────────────


@router.get("/overall-stats")
def overall_stats(admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    total_agents = db.query(func.count(User.id)).filter(User.role == "agent", User.is_active == True).scalar() or 0
    total_managers = db.query(func.count(User.id)).filter(User.role == "manager", User.is_active == True).scalar() or 0
    total_cb = db.query(func.count(CallBack.id)).scalar() or 0
    total_tr = db.query(func.count(Transfer.id)).scalar() or 0
    total_sa = db.query(func.count(Sale.id)).scalar() or 0
    total_opps = total_tr
    return OverallStats(
        totalAgents=total_agents,
        totalManagers=total_managers,
        totalCallbacks=total_cb,
        totalTransfers=total_tr,
        totalSales=total_sa,
        conversionRate=_safe_div(total_sa, total_opps),
    )


@router.get("/performance-overview")
def performance_overview(admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    managers_data = []
    managers = db.query(User).filter(User.role == "manager", User.is_active == True).all()
    for m in managers:
        a_ids = [a.id for a in db.query(User).filter(User.manager_id == m.id, User.is_active == True).all()]
        if a_ids:
            cb = db.query(func.count(CallBack.id)).filter(CallBack.employee_id.in_(a_ids)).scalar() or 0
            tr = db.query(func.count(Transfer.id)).filter(Transfer.employee_id.in_(a_ids)).scalar() or 0
            sa = db.query(func.count(Sale.id)).filter(Sale.employee_id.in_(a_ids)).scalar() or 0
        else:
            cb = tr = sa = 0
        managers_data.append(ManagerKpi(
            id=m.id, name=m.name, teamSize=len(a_ids),
            callbacks=cb, transfers=tr, sales=sa,
            conversionRate=_safe_div(sa, tr),
        ))

    agents_data = []
    agents = db.query(User).filter(User.role == "agent", User.is_active == True).all()
    mgr_names = {m.id: m.name for m in db.query(User).filter(User.role == "manager", User.is_active == True).all()}
    for a in agents:
        cb = db.query(func.count(CallBack.id)).filter(CallBack.employee_id == a.id).scalar() or 0
        tr = db.query(func.count(Transfer.id)).filter(Transfer.employee_id == a.id).scalar() or 0
        sa = db.query(func.count(Sale.id)).filter(Sale.employee_id == a.id).scalar() or 0
        agents_data.append(AgentKpi(
            id=a.id, name=a.name,
            managerName=mgr_names.get(a.manager_id, "Unassigned"),
            callbacks=cb, transfers=tr, sales=sa,
            conversionRate=_safe_div(sa, tr),
            isActive=1 if a.is_active else 0,
        ))

    agents_data.sort(key=lambda x: x.conversionRate, reverse=True)
    top_agents = agents_data[:10]
    bottom_agents = list(reversed(agents_data[-5:])) if len(agents_data) > 5 else list(reversed(agents_data))

    total_cb = sum(a.callbacks for a in agents_data)
    total_tr = sum(a.transfers for a in agents_data)
    total_sa = sum(a.sales for a in agents_data)

    return AdminPerformanceOverview(
        topManagers=sorted(managers_data, key=lambda x: x.conversionRate, reverse=True),
        topAgents=top_agents,
        bottomAgents=bottom_agents,
        totalCallbacks=total_cb,
        totalTransfers=total_tr,
        totalSales=total_sa,
        conversionRate=_safe_div(total_sa, total_tr),
    )


@router.get("/business-feed")
def business_feed(admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    user_map = {u.id: u.name for u in db.query(User).all()}
    events = []

    callbacks = db.query(CallBack).order_by(CallBack.created_at.desc()).limit(30).all()
    for cb in callbacks:
        events.append(BusinessFeedItem(
            type="callback", action="created",
            agentName=user_map.get(cb.employee_id, "Unknown"),
            description="created a callback",
            timestamp=cb.created_at.isoformat() if cb.created_at else None,
            id=cb.id,
        ))

    transfers = db.query(Transfer).order_by(Transfer.created_at.desc()).limit(30).all()
    for t in transfers:
        events.append(BusinessFeedItem(
            type="transfer", action="created",
            agentName=user_map.get(t.employee_id, "Unknown"),
            description="created a transfer",
            timestamp=t.created_at.isoformat() if t.created_at else None,
            id=t.id,
        ))

    sales = db.query(Sale).order_by(Sale.created_at.desc()).limit(30).all()
    for s in sales:
        events.append(BusinessFeedItem(
            type="sale", action="closed",
            agentName=user_map.get(s.employee_id, "Unknown"),
            description="closed a sale",
            timestamp=s.created_at.isoformat() if s.created_at else None,
            id=s.id,
        ))

    events.sort(key=lambda e: e.timestamp or "", reverse=True)
    return events[:50]


@router.get("/callbacks")
def admin_callbacks(
    status: str | None = Query(None),
    from_date: date | None = Query(None, description="Filter by start date (YYYY-MM-DD)"),
    to_date: date | None = Query(None, description="Filter by end date (YYYY-MM-DD)"),
    employee_id: int | None = Query(None, description="Filter by employee"),
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(20, ge=1, le=500, description="Items per page"),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    q = db.query(CallBack)
    if status:
        q = q.filter(CallBack.status == status)
    if from_date:
        from_dt = datetime.combine(from_date, datetime.min.time())
        q = q.filter(CallBack.created_at >= from_dt)
    if to_date:
        to_dt = datetime.combine(to_date, datetime.max.time())
        q = q.filter(CallBack.created_at <= to_dt)
    if employee_id:
        q = q.filter(CallBack.employee_id == employee_id)
    total = q.count()
    total_pages = max(1, (total + per_page - 1) // per_page)
    skip = (page - 1) * per_page
    items = [_build_callback_out(c, c.customer) for c in q.order_by(CallBack.created_at.desc()).offset(skip).limit(per_page).all()]
    return {"items": items, "total": total, "page": page, "totalPages": total_pages}


@router.get("/callbacks/export")
def admin_callbacks_export(
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
    from_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    to_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    employee_id: int | None = Query(None, description="Filter by employee"),
):
    from_dt = datetime.combine(from_date, datetime.min.time())
    to_dt = datetime.combine(to_date, datetime.max.time())
    q = db.query(CallBack)
    q = q.filter(CallBack.scheduled_datetime >= from_dt, CallBack.scheduled_datetime <= to_dt)
    if employee_id:
        q = q.filter(CallBack.employee_id == employee_id)
    rows = q.order_by(CallBack.scheduled_datetime.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Callbacks Export"

    headers = [
        "ID", "Agent", "Agent Email", "Business Name", "Owner Name",
        "Business Phone", "Owner Phone", "Email", "Business Address", "Postcode",
        "Scheduled Date", "Day of Week", "Status", "Outcome",
        "Account Number", "MPAN", "MPRN", "MSN",
        "Notes", "Not Interested Reason",
        "Electric Supplier", "Elec Contract Length", "Elec Meter Type",
        "Elec Commission Type", "Elec Day Rate", "Elec Night Rate",
        "Elec Evening Rate", "Elec Standing Rate",
        "Elec Non-Com Day", "Elec Non-Com Night", "Elec Non-Com Evening", "Elec Non-Com Standing",
        "Elec Broker Charge",
        "Gas Supplier", "Gas Contract Length",
        "Gas Unit Rate", "Gas Standing Rate",
        "Gas Non-Com Unit Rate", "Gas Non-Com Standing",
        "Gas Broker Charge",
        "Created At",
    ]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, cb in enumerate(rows, 2):
        c = cb.customer
        values = [
            cb.id,
            cb.employee.name if cb.employee else "",
            cb.employee.email if cb.employee else "",
            c.business_name if c else "",
            c.owner_name if c else "",
            c.business_phone if c else "",
            c.owner_phone if c else "",
            c.email if c else "",
            c.business_address if c else "",
            c.postcode if c else "",
            cb.scheduled_datetime.strftime("%Y-%m-%d %H:%M") if cb.scheduled_datetime else "",
            cb.scheduled_datetime.strftime("%A") if cb.scheduled_datetime else "",
            cb.status or "",
            cb.outcome or "",
            cb.account_number or "",
            cb.mpan or "",
            cb.mprn or "",
            cb.msn or "",
            cb.notes or "",
            cb.not_interested_reason or "",
            cb.elec_offer_supplier or "",
            cb.elec_offer_contract_length or "",
            cb.elec_offer_meter_type or "",
            cb.elec_offer_commission_type or "",
            cb.elec_commission_day_rate or "",
            cb.elec_commission_night_rate or "",
            cb.elec_commission_evening_rate or "",
            cb.elec_commission_standing or "",
            cb.elec_noncom_day_rate or "",
            cb.elec_noncom_night_rate or "",
            cb.elec_noncom_evening_rate or "",
            cb.elec_noncom_standing or "",
            cb.elec_broker_charge or "",
            cb.gas_offer_supplier or "",
            cb.gas_offer_contract_length or "",
            cb.gas_commission_unit_rate or "",
            cb.gas_commission_standing or "",
            cb.gas_noncom_unit_rate or "",
            cb.gas_noncom_standing or "",
            cb.gas_broker_charge or "",
            cb.created_at.strftime("%Y-%m-%d %H:%M") if cb.created_at else "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    col_widths = [6, 20, 28, 22, 18, 16, 16, 28, 28, 12, 18, 14, 12, 14, 16, 16, 16, 16, 30, 22, 18, 18, 14, 18, 12, 12, 14, 14, 14, 14, 16, 14, 14, 18, 18, 14, 14, 16, 16, 16, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"callbacks_{from_date.isoformat()}_{to_date.isoformat()}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/transfers")
def admin_transfers(
    status: str | None = Query(None),
    from_date: date | None = Query(None, description="Filter by start date (YYYY-MM-DD)"),
    to_date: date | None = Query(None, description="Filter by end date (YYYY-MM-DD)"),
    employee_id: int | None = Query(None, description="Filter by employee"),
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(20, ge=1, le=500, description="Items per page"),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    q = db.query(Transfer)
    if status:
        q = q.filter(Transfer.status == status)
    if from_date:
        from_dt = datetime.combine(from_date, datetime.min.time())
        q = q.filter(Transfer.created_at >= from_dt)
    if to_date:
        to_dt = datetime.combine(to_date, datetime.max.time())
        q = q.filter(Transfer.created_at <= to_dt)
    if employee_id:
        q = q.filter(Transfer.employee_id == employee_id)
    total = q.count()
    total_pages = max(1, (total + per_page - 1) // per_page)
    skip = (page - 1) * per_page
    items = [_transfer_out(t, t.customer) for t in q.order_by(Transfer.created_at.desc()).offset(skip).limit(per_page).all()]
    return {"items": items, "total": total, "page": page, "totalPages": total_pages}


@router.get("/transfers/export")
def admin_transfers_export(
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
    from_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    to_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    employee_id: int | None = Query(None, description="Filter by employee"),
):
    from_dt = datetime.combine(from_date, datetime.min.time())
    to_dt = datetime.combine(to_date, datetime.max.time())
    q = db.query(Transfer)
    q = q.filter(Transfer.created_at >= from_dt, Transfer.created_at <= to_dt)
    if employee_id:
        q = q.filter(Transfer.employee_id == employee_id)
    rows = q.order_by(Transfer.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transfers Export"

    headers = [
        "ID", "Agent", "Agent Email", "Business Name", "Owner Name",
        "Business Phone", "Owner Phone", "Email", "Business Address", "Postcode",
        "Utility Type", "Status", "Outcome", "Scheduled Date",
        "Account Number", "MPAN", "MPRN", "MSN",
        "Notes", "Not Interested Reason",
        "Electric Supplier", "Elec Contract Length", "Elec Meter Type",
        "Elec Commission Type", "Elec Day Rate", "Elec Night Rate",
        "Elec Evening Rate", "Elec Standing Rate",
        "Elec Non-Com Day", "Elec Non-Com Night", "Elec Non-Com Evening", "Elec Non-Com Standing",
        "Elec Broker Charge",
        "Gas Supplier", "Gas Contract Length",
        "Gas Unit Rate", "Gas Standing Rate",
        "Gas Non-Com Unit Rate", "Gas Non-Com Standing",
        "Gas Broker Charge",
        "Created At",
    ]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, t in enumerate(rows, 2):
        c = t.customer
        values = [
            t.id,
            t.employee.name if t.employee else "",
            t.employee.email if t.employee else "",
            c.business_name if c else "",
            c.owner_name if c else "",
            c.business_phone if c else "",
            c.owner_phone if c else "",
            c.email if c else "",
            c.business_address if c else "",
            c.postcode if c else "",
            t.utility_type or "",
            t.status or "",
            t.outcome or "",
            t.scheduled_datetime.strftime("%Y-%m-%d %H:%M") if t.scheduled_datetime else "",
            t.account_number or "",
            t.mpan or "",
            t.mprn or "",
            t.msn or "",
            t.notes or "",
            t.not_interested_reason or "",
            t.elec_offer_supplier or "",
            t.elec_offer_contract_length or "",
            t.elec_offer_meter_type or "",
            t.elec_offer_commission_type or "",
            t.elec_commission_day_rate or "",
            t.elec_commission_night_rate or "",
            t.elec_commission_evening_rate or "",
            t.elec_commission_standing or "",
            t.elec_noncom_day_rate or "",
            t.elec_noncom_night_rate or "",
            t.elec_noncom_evening_rate or "",
            t.elec_noncom_standing or "",
            t.elec_broker_charge or "",
            t.gas_offer_supplier or "",
            t.gas_offer_contract_length or "",
            t.gas_commission_unit_rate or "",
            t.gas_commission_standing or "",
            t.gas_noncom_unit_rate or "",
            t.gas_noncom_standing or "",
            t.gas_broker_charge or "",
            t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    col_widths = [6, 20, 28, 22, 18, 16, 16, 28, 28, 12, 14, 12, 14, 18, 16, 16, 16, 16, 30, 22, 18, 18, 14, 18, 12, 12, 14, 14, 14, 14, 16, 14, 14, 18, 18, 14, 14, 16, 16, 16, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"transfers_{from_date.isoformat()}_{to_date.isoformat()}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/sales")
def admin_sales(
    status: str | None = Query(None),
    from_date: date | None = Query(None, description="Filter by start date (YYYY-MM-DD)"),
    to_date: date | None = Query(None, description="Filter by end date (YYYY-MM-DD)"),
    employee_id: int | None = Query(None, description="Filter by employee"),
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(20, ge=1, le=500, description="Items per page"),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    q = db.query(Sale)
    if status:
        q = q.filter(Sale.cot_status == status)
    if from_date:
        from_dt = datetime.combine(from_date, datetime.min.time())
        q = q.filter(Sale.created_at >= from_dt)
    if to_date:
        to_dt = datetime.combine(to_date, datetime.max.time())
        q = q.filter(Sale.created_at <= to_dt)
    if employee_id:
        q = q.filter(Sale.employee_id == employee_id)
    total = q.count()
    total_pages = max(1, (total + per_page - 1) // per_page)
    skip = (page - 1) * per_page
    items = [_sale_out(s, s.customer) for s in q.order_by(Sale.created_at.desc()).offset(skip).limit(per_page).all()]
    return {"items": items, "total": total, "page": page, "totalPages": total_pages}


@router.get("/sales/export")
def admin_sales_export(
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
    from_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    to_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    employee_id: int | None = Query(None, description="Filter by employee"),
):
    from_dt = datetime.combine(from_date, datetime.min.time())
    to_dt = datetime.combine(to_date, datetime.max.time())
    q = db.query(Sale)
    q = q.filter(Sale.created_at >= from_dt, Sale.created_at <= to_dt)
    if employee_id:
        q = q.filter(Sale.employee_id == employee_id)
    rows = q.order_by(Sale.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Export"

    headers = [
        "ID", "Agent", "Agent Email", "Business Name", "Owner Name",
        "Business Phone", "Owner Phone", "Email", "Business Address", "Postcode",
        "Owner Full Name", "Home Address", "Date of Birth", "Business Type",
        "Bill Frequency", "Payment Method", "Bank Name", "Account Type",
        "Account Title", "Sort Code", "Bank Account Number",
        "COT Status", "COT Date", "COT Notes", "Commission Amount",
        "Notes",
        "Created At",
    ]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, s in enumerate(rows, 2):
        c = s.customer
        values = [
            s.id,
            s.employee.name if s.employee else "",
            s.employee.email if s.employee else "",
            c.business_name if c else "",
            c.owner_name if c else "",
            c.business_phone if c else "",
            c.owner_phone if c else "",
            c.email if c else "",
            c.business_address if c else "",
            c.postcode if c else "",
            s.owner_full_name or "",
            s.home_address or "",
            s.date_of_birth.isoformat() if s.date_of_birth else "",
            s.business_type or "",
            s.bill_frequency or "",
            s.payment_method or "",
            s.bank_name or "",
            s.account_type or "",
            s.account_title or "",
            s.sort_code or "",
            s.bank_account_number or "",
            s.cot_status or "",
            s.cot_date.isoformat() if s.cot_date else "",
            s.cot_notes or "",
            s.commission_amount or 0,
            s.notes or "",
            s.created_at.strftime("%Y-%m-%d %H:%M") if s.created_at else "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    col_widths = [6, 20, 28, 22, 18, 16, 16, 28, 28, 12, 22, 28, 14, 14, 14, 16, 18, 14, 22, 14, 20, 14, 14, 24, 16, 30, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"sales_{from_date.isoformat()}_{to_date.isoformat()}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/sales/{sale_id}")
def admin_sale_detail(sale_id: int, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    sale = db.query(Sale).filter(Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    return _sale_out(sale, sale.customer)


@router.get("/callbacks/{callback_id}")
def admin_callback_detail(callback_id: int, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    cb = db.query(CallBack).filter(CallBack.id == callback_id).first()
    if not cb:
        raise HTTPException(status_code=404, detail="Callback not found")
    return _build_callback_out(cb, cb.customer)


@router.get("/audit-log")
def get_audit_log(admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    logs = db.query(ActivityLog).order_by(ActivityLog.created_at.desc()).limit(100).all()
    return [
        {
            "id": log.id,
            "userId": log.user_id,
            "action": log.action,
            "entityType": log.entity_type,
            "entityId": log.entity_id,
            "description": log.description,
            "ipAddress": log.ip_address,
            "createdAt": log.created_at.isoformat() if log.created_at else None,
        }
        for log in logs
    ]


@router.get("/salary/slip/{user_id}")
def admin_salary_slip(
    user_id: int,
    month: int = Query(None, ge=1, le=12),
    year: int = Query(None, ge=2020, le=2100),
    commission: float = Query(0.0, ge=0),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    agent = db.query(User).filter(User.id == user_id).first()
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")

    now = datetime.now()
    m = month or now.month
    y = year or now.year

    working_days, present_days, absent_days = _attendance_summary(db, agent.id, m, y)
    base = agent.monthly_salary or 0

    basic      = int(round(base * 0.50))
    hra        = int(round(base * 0.15))
    utility    = int(round(base * 0.30))
    conveyance = int(round(base * 0.05))
    comm       = int(commission)
    gross      = basic + hra + utility + conveyance + comm

    daily_rate        = (basic + hra + utility + conveyance) / working_days if working_days > 0 else 0
    absent_deduction  = int(round(daily_rate * absent_days))
    total_deductions  = absent_deduction
    net_salary        = gross

    earnings_rows = [
        ("Basic Salary",           _fmt(basic),      False),
        ("House Rent Allowance",   _fmt(hra),        False),
        ("Utility Allowance",      _fmt(utility),    False),
        ("Conveyance Allowance",   _fmt(conveyance), False),
    ]
    if comm:
        earnings_rows.append(("Commission", _fmt(comm), False))
    earnings_rows.append(("Gross Salary", _fmt(gross), True))

    deductions_rows = [
        ("Absent Days Deduction", _fmt(absent_deduction), False),
        ("Total Deductions",      _fmt(total_deductions), True),
    ]

    pdf = _SalaryPDF()
    pdf.build_slip(
        period        = f"for the month of {_month_name(m)} {y}",
        employee_name = agent.name,
        employee_id   = _employee_id(agent),
        designation   = agent.designation   or "—",
        department    = agent.department    or "—",
        cnic          = agent.cnic          or "—",
        doj           = agent.date_of_joining.strftime("%d/%m/%Y") if agent.date_of_joining else "—",
        working_days  = str(working_days),
        present_days  = str(present_days),
        absent_days   = str(absent_days),
        earnings      = earnings_rows,
        deductions    = deductions_rows,
        net_salary    = net_salary,
        generated_at  = datetime.now().strftime("%B %d, %Y at %I:%M %p"),
    )

    return Response(
        content=pdf.output(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="Salary_Slip_{_month_name(m)}_{y}_{agent.name}.pdf"'},
    )


@router.post("/send-otp")
def admin_send_otp(request: SendOTPRequest, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    email = request.email.strip().lower()
    otp = generate_otp()
    expiry = datetime.utcnow() + timedelta(minutes=5)

    existing = db.query(StaffOTP).filter(StaffOTP.email == email).first()
    if existing:
        existing.otp_code = otp
        existing.otp_expiry = expiry
        existing.is_verified = False
    else:
        db.add(StaffOTP(email=email, otp_code=otp, otp_expiry=expiry))

    db.commit()

    sent = send_otp_email(email, otp)
    return {"message": "If this email exists, an OTP has been sent.", "sent": sent, "otp": otp if not sent else None}


@router.post("/verify-otp")
def admin_verify_otp(request: VerifyOTPRequest, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    email = request.email.strip().lower()
    record = db.query(StaffOTP).filter(StaffOTP.email == email, StaffOTP.is_verified == False).first()
    if not record:
        raise HTTPException(status_code=400, detail="No OTP requested. Please request a new OTP.")
    if record.otp_expiry < datetime.utcnow():
        raise HTTPException(status_code=400, detail="OTP has expired. Please request a new OTP.")
    if record.otp_code != request.otp:
        raise HTTPException(status_code=400, detail="Invalid OTP. Please try again.")

    record.is_verified = True
    record.otp_code = None
    record.otp_expiry = None
    db.commit()

    return {"message": "Email verified successfully", "verified": True}
