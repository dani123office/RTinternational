import io
from calendar import monthrange
from datetime import datetime, date, timedelta
from fastapi import APIRouter, Depends, HTTPException, Query, Request, Response
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from ..database import get_db
from ..models import User, Attendance
from .auth import get_current_user
from pydantic import BaseModel
from ..schemas import AttendanceCheckIn, AttendanceCheckOut, AttendanceOut, AttendanceSummary, UserAttendanceToday, LateArrivalReport
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class AttendanceRecordOut(BaseModel):
    record: AttendanceOut
    agentName: str
    agentEmail: str


class AttendanceFeedItem(BaseModel):
    id: int
    userId: int
    userName: str
    userEmail: str
    date: date
    checkIn: Optional[datetime] = None
    checkOut: Optional[datetime] = None
    status: str
    checkin_reason: Optional[str] = None
    checkout_reason: Optional[str] = None
    expected_arrival_time: Optional[str] = None
    late_arrival_reason: Optional[str] = None
from ..utils.logger import log_activity, get_client_ip

router = APIRouter(prefix="/api/attendance", tags=["attendance"])

PKT_OFFSET = timedelta(hours=5)

# Weekday-aware thresholds: Mon-Thu -> 14:10, Fri -> 15:10 (PKT)
LATE_THRESHOLD_MON_THU = (14, 10)
LATE_THRESHOLD_FRIDAY = (15, 10)

def _now_pkt():
    return datetime.utcnow() + PKT_OFFSET

def _today_pkt():
    return _now_pkt().date()

def _attendance_to_out(a: Attendance) -> AttendanceOut:
    return AttendanceOut(
        id=a.id,
        userId=a.user_id,
        date=a.date,
        checkIn=a.check_in,
        checkOut=a.check_out,
        status=a.status,
        checkout_reason=a.checkout_reason,
        checkin_reason=a.checkin_reason,
        expected_arrival_time=a.expected_arrival_time.strftime("%H:%M") if a.expected_arrival_time else None,
        late_arrival_reason=a.late_arrival_reason,
        createdAt=a.created_at,
        updatedAt=a.updated_at,
    )


@router.get("/today")
def get_today_attendance(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    now = _now_pkt()
    today = now.date()
    record = db.query(Attendance).filter(
        Attendance.user_id == current_user.id,
        Attendance.date == today,
    ).first()

    if record and record.check_in and not record.check_out and now.hour >= 22:
        record.check_out = now
        record.checkout_reason = (record.checkout_reason or "") + (" | " if record.checkout_reason else "") + "Auto check-out at 10 PM"
        record.updated_at = now
        db.commit()
        db.refresh(record)

    if record:
        return _attendance_to_out(record)
    return None


@router.post("/check-in")
def check_in(
    dto: AttendanceCheckIn,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    now = _now_pkt()
    today = now.date()

    existing = db.query(Attendance).filter(
        Attendance.user_id == current_user.id,
        Attendance.date == today,
    ).first()
    if existing and existing.check_in:
        raise HTTPException(status_code=400, detail="Already checked in today")

    # determine weekday-specific threshold
    wd = now.weekday()  # 0=Mon ... 4=Fri
    if wd == 4:
        th_h, th_m = LATE_THRESHOLD_FRIDAY
    else:
        th_h, th_m = LATE_THRESHOLD_MON_THU

    is_late = now.hour > th_h or (now.hour == th_h and now.minute > th_m)
    status = "late" if is_late else "present"

    if existing:
        existing.check_in = now
        existing.status = status
        if dto.checkin_reason is not None:
            existing.checkin_reason = dto.checkin_reason
        existing.updated_at = now
        db.commit()
        db.refresh(existing)
        log_activity(db, current_user.id, "checked_in", "attendance", existing.id,
                     f"Checked in at {now.strftime('%H:%M')}",
                     get_client_ip(request))
        return _attendance_to_out(existing)

    record = Attendance(
        user_id=current_user.id,
        date=today,
        check_in=now,
        status=status,
        checkin_reason=dto.checkin_reason,
        created_at=now,
        updated_at=now,
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    log_activity(db, current_user.id, "checked_in", "attendance", record.id,
                 f"Checked in at {now.strftime('%H:%M')}",
                 get_client_ip(request))
    return _attendance_to_out(record)


@router.post("/check-out")
def check_out(
    dto: AttendanceCheckOut,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    now = _now_pkt()
    today = now.date()

    record = db.query(Attendance).filter(
        Attendance.user_id == current_user.id,
        Attendance.date == today,
    ).first()
    if not record:
        raise HTTPException(status_code=400, detail="No check-in found for today")
    if record.check_out:
        raise HTTPException(status_code=400, detail="Already checked out today")

    record.check_out = now
    if dto.checkout_reason is not None:
        record.checkout_reason = (record.checkout_reason or "") + (" | " if record.checkout_reason else "") + dto.checkout_reason
    record.updated_at = now
    db.commit()
    db.refresh(record)
    log_activity(db, current_user.id, "checked_out", "attendance", record.id,
                 f"Checked out at {now.strftime('%H:%M')}",
                 get_client_ip(request))
    return _attendance_to_out(record)


@router.post("/report-late-arrival")
def report_late_arrival(
    dto: LateArrivalReport,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if dto.date > _today_pkt():
        raise HTTPException(status_code=400, detail="Cannot report late arrival for a future date")

    record = db.query(Attendance).filter(
        Attendance.user_id == current_user.id,
        Attendance.date == dto.date,
    ).first()

    if not record:
        record = Attendance(
            user_id=current_user.id,
            date=dto.date,
            status="late",
            check_in=None,
            check_out=None,
            expected_arrival_time=dto.expected_arrival_time,
            late_arrival_reason=dto.reason,
            created_at=_now_pkt(),
            updated_at=_now_pkt(),
        )
        db.add(record)
    else:
        record.expected_arrival_time = dto.expected_arrival_time
        record.late_arrival_reason = dto.reason
        if record.status == "present":
            record.status = "late"
        record.updated_at = _now_pkt()

    db.commit()
    db.refresh(record)
    log_activity(db, current_user.id, "reported_late", "attendance", record.id,
                 f"Reported late arrival for {dto.date}: expected {dto.expected_arrival_time.strftime('%H:%M')}",
                 get_client_ip(request))
    return _attendance_to_out(record)


@router.get("/my-history")
def my_history(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
):
    total = db.query(func.count(Attendance.id)).filter(
        Attendance.user_id == current_user.id,
    ).scalar()
    records = db.query(Attendance).filter(
        Attendance.user_id == current_user.id,
    ).order_by(
        Attendance.date.desc()
    ).offset((page - 1) * per_page).limit(per_page).all()

    return {
        "items": [_attendance_to_out(r) for r in records],
        "total": total,
        "page": page,
        "perPage": per_page,
        "totalPages": (total + per_page - 1) // per_page,
    }


@router.get("/stats")
def my_stats(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    month: int = Query(None),
    year: int = Query(None),
):
    today = _today_pkt()
    m = month or today.month
    y = year or today.year
    _, last_day = monthrange(y, m)
    start = date(y, m, 1)
    end = date(y, m, last_day)

    # Don't count future working days for current month
    if m == today.month and y == today.year:
        effective_end = min(end, today)
    else:
        effective_end = end

    # Start counting from first attendance record (or month start, whichever is later)
    first_record = db.query(func.min(Attendance.date)).filter(
        Attendance.user_id == current_user.id,
        Attendance.date >= start,
    ).scalar()
    effective_start = max(start, first_record) if first_record else start

    total_working = sum(
        1 for d in range((effective_end - effective_start).days + 1)
        if (effective_start + timedelta(days=d)).weekday() < 5
    )

    present = db.query(func.count(Attendance.id)).filter(
        Attendance.user_id == current_user.id,
        Attendance.date >= start,
        Attendance.date <= end,
        Attendance.status == "present",
    ).scalar() or 0

    late = db.query(func.count(Attendance.id)).filter(
        Attendance.user_id == current_user.id,
        Attendance.date >= start,
        Attendance.date <= end,
        Attendance.status == "late",
    ).scalar() or 0

    absent = max(0, total_working - present - late)

    return AttendanceSummary(
        presentCount=present,
        lateCount=late,
        absentCount=absent,
        totalDays=total_working,
    )


@router.get("/team-today")
def team_today(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in ("manager", "admin"):
        raise HTTPException(status_code=403, detail="Only managers and admins can view team attendance")
    today = _today_pkt()

    if current_user.role == "manager":
        team = db.query(User).filter(
            User.manager_id == current_user.id,
            User.role == "agent",
        ).all()
    else:
        team = db.query(User).filter(User.role == "agent").all()

    result = []
    for agent in team:
        record = db.query(Attendance).filter(
            Attendance.user_id == agent.id,
            Attendance.date == today,
        ).first()
        result.append(UserAttendanceToday(
            userId=agent.id,
            userName=agent.name,
            userEmail=agent.email,
            attendance=_attendance_to_out(record) if record else None,
        ))

    return result


@router.get("/agent-history/{agent_id}")
def agent_history(
    agent_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
):
    if current_user.role not in ("manager", "admin"):
        raise HTTPException(status_code=403, detail="Only managers and admins can view agent attendance")

    agent = db.query(User).filter(User.id == agent_id).first()
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")

    if current_user.role == "manager" and agent.manager_id != current_user.id:
        raise HTTPException(status_code=403, detail="You can only view your own team members")

    total = db.query(func.count(Attendance.id)).filter(
        Attendance.user_id == agent_id,
    ).scalar()
    records = db.query(Attendance).filter(
        Attendance.user_id == agent_id,
    ).order_by(
        Attendance.date.desc()
    ).offset((page - 1) * per_page).limit(per_page).all()

    return {
        "items": [_attendance_to_out(r) for r in records],
        "total": total,
        "page": page,
        "perPage": per_page,
        "totalPages": (total + per_page - 1) // per_page,
    }


@router.get("/record/{record_id}")
def attendance_record(
    record_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in ("manager", "admin"):
        raise HTTPException(status_code=403, detail="Only managers and admins can view attendance records")

    record = db.query(Attendance).filter(Attendance.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Attendance record not found")

    agent = db.query(User).filter(User.id == record.user_id).first()
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")

    if current_user.role == "manager" and agent.manager_id != current_user.id:
        raise HTTPException(status_code=403, detail="You can only view your own team members")

    return AttendanceRecordOut(
        record=_attendance_to_out(record),
        agentName=agent.name,
        agentEmail=agent.email,
    )


@router.get("/feed")
def attendance_feed(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    from_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    to_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    employee_id: Optional[int] = Query(None, description="Filter by employee"),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
):
    if current_user.role not in ("manager", "admin"):
        raise HTTPException(status_code=403, detail="Only managers and admins can view attendance feed")

    query = db.query(Attendance, User).join(User, Attendance.user_id == User.id)

    if current_user.role == "manager":
        query = query.filter(User.manager_id == current_user.id, User.role == "agent")

    query = query.filter(Attendance.date >= from_date, Attendance.date <= to_date)

    if employee_id:
        query = query.filter(Attendance.user_id == employee_id)

    total = query.with_entities(func.count()).scalar()
    rows = query.order_by(Attendance.date.desc(), User.name.asc()).offset(
        (page - 1) * per_page
    ).limit(per_page).all()

    items = [
        AttendanceFeedItem(
            id=a.id,
            userId=a.user_id,
            userName=u.name,
            userEmail=u.email,
            date=a.date,
            checkIn=a.check_in,
            checkOut=a.check_out,
            status=a.status,
            checkin_reason=a.checkin_reason,
            checkout_reason=a.checkout_reason,
            expected_arrival_time=a.expected_arrival_time.strftime("%H:%M") if a.expected_arrival_time else None,
            late_arrival_reason=a.late_arrival_reason,
        )
        for a, u in rows
    ]

    return {
        "items": items,
        "total": total,
        "page": page,
        "perPage": per_page,
        "totalPages": (total + per_page - 1) // per_page if total else 0,
    }


@router.get("/export")
def attendance_export(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    from_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    to_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    employee_id: Optional[int] = Query(None, description="Filter by employee"),
):
    if current_user.role not in ("manager", "admin"):
        raise HTTPException(status_code=403, detail="Only managers and admins can export attendance")

    query = db.query(Attendance, User).join(User, Attendance.user_id == User.id)

    if current_user.role == "manager":
        query = query.filter(User.manager_id == current_user.id, User.role == "agent")

    query = query.filter(Attendance.date >= from_date, Attendance.date <= to_date)

    if employee_id:
        query = query.filter(Attendance.user_id == employee_id)

    rows = query.order_by(Attendance.date.desc(), User.name.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Feed"

    headers = ["Date", "Employee Name", "Email", "Status", "Check In", "Check Out", "Check-in Reason", "Check-out Reason", "Exp. Arrival Time", "Late Arrival Reason"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, (a, u) in enumerate(rows, 2):
        values = [
            a.date.isoformat() if a.date else "",
            u.name or "",
            u.email or "",
            a.status.capitalize() if a.status else "",
            a.check_in.strftime("%H:%M") if a.check_in else "",
            a.check_out.strftime("%H:%M") if a.check_out else "",
            a.checkin_reason or "",
            a.checkout_reason or "",
            a.expected_arrival_time.strftime("%H:%M") if a.expected_arrival_time else "",
            a.late_arrival_reason or "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["H"].width = 28
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 28

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"attendance_{from_date.isoformat()}_{to_date.isoformat()}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
